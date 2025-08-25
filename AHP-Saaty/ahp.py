from __future__ import annotations
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Optional

RI_TABLE = {1:0,2:0,3:0.58,4:0.90,5:1.12,6:1.24,7:1.32,8:1.41,9:1.45,10:1.49}

DEFAULT_RANGES = [
    {"label":"Bajo","min":1.00,"max":1.80},
    {"label":"Medio Bajo","min":1.81,"max":2.60},
    {"label":"Medio","min":2.61,"max":3.40},
    {"label":"Medio Alto","min":3.41,"max":4.20},
    {"label":"Alto","min":4.21,"max":4.60},
    {"label":"Extremo","min":4.61,"max":5.00},
]

def _is_pairwise_matrix(arr: np.ndarray) -> bool:
    n, m = arr.shape
    if n != m or n < 3: 
        return False
    # diagonal ~ 1
    if not np.allclose(np.diag(arr), np.ones(n), atol=1e-3, rtol=1e-3):
        return False
    # positiva
    if np.any(arr <= 0):
        return False
    # reciprocidad
    for i in range(n):
        for j in range(n):
            if i == j: 
                continue
            if not np.isfinite(arr[i, j]) or not np.isfinite(arr[j, i]):
                return False
            if not np.isclose(arr[i, j] * arr[j, i], 1.0, rtol=1e-2, atol=1e-2):
                return False
    return True

def parse_excel(path: str) -> Tuple[List[str], np.ndarray]:
    """Lee un Excel sin pandas: intenta detectar criterios y matriz pareada
       usando openpyxl directamente.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)

    def sheet_to_matrix(ws):
        # Convierte hoja en lista de listas, recortando filas/cols vacías
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        # elimina filas y columnas completamente vacías
        # recorta nulos al borde
        # filas vacías:
        data = [r for r in data if any(c is not None and str(c).strip() != "" for c in r)]
        if not data:
            return []
        # columnas vacías:
        # determina qué columnas tienen algún valor
        max_len = max(len(r) for r in data)
        cols_used = [False]*max_len
        for r in data:
            for j, c in enumerate(r + [None]*(max_len-len(r))):
                if c is not None and str(c).strip() != "":
                    cols_used[j] = True
        # recorta
        out = []
        for r in data:
            row = []
            for j, c in enumerate(r + [None]*(max_len-len(r))):
                if cols_used[j]:
                    row.append(c)
            out.append(row)
        return out

    def is_pairwise(arr: np.ndarray) -> bool:
        n, m = arr.shape
        if n != m or n < 3:
            return False
        if not np.allclose(np.diag(arr), np.ones(n), atol=1e-3, rtol=1e-3):
            return False
        if np.any(arr <= 0):
            return False
        for i in range(n):
            for j in range(n):
                if i == j: 
                    continue
                if not np.isfinite(arr[i, j]) or not np.isfinite(arr[j, i]):
                    return False
                if not np.isclose(arr[i, j] * arr[j, i], 1.0, rtol=1e-2, atol=1e-2):
                    return False
        return True

    # Caso 1: hojas con nombre standard
    ws_crit = wb["criterios"] if "criterios" in wb.sheetnames else None
    ws_mat  = wb["matriz"]    if "matriz"    in wb.sheetnames else None
    if ws_crit and ws_mat:
        crit = []
        for cell in ws_crit["A"]:
            if cell.value is None:
                continue
            txt = str(cell.value).strip()
            if txt and txt.lower() not in ("criterio", "criterios"):
                crit.append(txt)
        n = len(crit)

        mdat = sheet_to_matrix(ws_mat)
        # intenta subbloques n x n en primeras 3 filas/cols
        for r0 in range(min(3, max(1, len(mdat)-n+1))):
            for c0 in range(min(3, max(1, len(mdat[0])-n+1))):
                block = []
                ok = True
                for i in range(n):
                    if r0+i >= len(mdat):
                        ok = False; break
                    row = mdat[r0+i]
                    if c0+n > len(row):
                        ok = False; break
                    block.append(row[c0:c0+n])
                if not ok:
                    continue
                try:
                    arr = np.array(block, dtype=float)
                except Exception:
                    continue
                if is_pairwise(arr):
                    return crit, arr

    # Caso 2: autodetección en cualquier hoja
    for name in wb.sheetnames:
        ws = wb[name]
        mdat = sheet_to_matrix(ws)
        if not mdat or not mdat[0]:
            continue
        R, C = len(mdat), len(mdat[0])
        for n in range(min(R, C), 2, -1):
            for r0 in range(0, R-n+1):
                for c0 in range(0, C-n+1):
                    block = [mdat[r0+i][c0:c0+n] for i in range(n)]
                    try:
                        arr = np.array(block, dtype=float)
                    except Exception:
                        continue
                    if is_pairwise(arr):
                        crit = [f"Criterio {i+1}" for i in range(n)]
                        return crit, arr
    raise ValueError("No se detectó una matriz pareada válida en el Excel.")

def ahp_weights(pairwise: np.ndarray) -> Dict:
    n = pairwise.shape[0]
    col_sums = pairwise.sum(axis=0)
    norm = pairwise / col_sums
    w = norm.mean(axis=1)
    w = w / w.sum()  # normaliza
    # lambda max
    Aw = pairwise.dot(w)
    lambdas = Aw / w
    lambda_max = lambdas.mean()
    CI = (lambda_max - n) / (n - 1) if n > 1 else 0.0
    RI = RI_TABLE.get(n, 1.49)
    CR = 0.0 if RI == 0 else CI / RI
    return {
        "weights": w.tolist(),
        "lambda_max": float(lambda_max),
        "CI": float(CI),
        "CR": float(CR),
        "valid": CR < 0.10
    }

def score_ic(weights: List[float], scores: Dict[str, float], criteria: List[str]) -> float:
    if len(weights) != len(criteria):
        raise ValueError("Dimensión de weights no coincide con criterios.")
    vec = [float(scores.get(c, 3)) for c in criteria]
    return float(np.dot(weights, np.array(vec)))

def categorize(IC: float, ranges: List[Dict]) -> str:
    for r in ranges:
        if r["min"] <= IC <= r["max"]:
            return r["label"]
    return "No clasificado"
